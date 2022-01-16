from abc import ABC, abstractmethod
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import logging
import decimal

logger = logging.getLogger(__name__)

class Import_data(ABC):

    @abstractmethod
    def process_import(self):
        pass


class Excel_import(Import_data):
    """import data from excel file using openpyxl"""

    def __init__(self, filename, year):
        self.file = filename
        self.year = year
        self.year_pre = str(year)[2:]

    def process_import(self, model):
        try:
            book = openpyxl.load_workbook(self.file)
            plants_sheet = book['PLNT'+self.year_pre]
            generators_sheet = book['GEN'+self.year_pre]
            plant_model = model.Plant
            plant_info_model = model.Plant_information
            energy_model = model.Energy
            self.save_plants(plants_sheet, plant_model, plant_info_model)
            self.save_energy(generators_sheet, plant_model, plant_info_model, energy_model)
            logger.warning('import data done')
            return True
        except KeyError as E:
            logger.exception(E)
            return False      

    def save_plants(self, plants_sheet, plant_model, plant_info_model):
        for row in plants_sheet.iter_rows(min_row=3):
            name = row[2].value
            facility_code = row[3].value
            state = row[1].value
            latitude = row[17].value
            longitude = row[18].value
            primary_fuel = row[21].value
            num_generator = row[20].value
            nameplate_capacity = row[25].value
            num_boilers = row[19].value
            year = self.year
            try:
                plant = plant_model.objects.create(name=name, facility_code=facility_code,
                                                   state=state, latitude=latitude,
                                                   longitude=longitude)
                plant_info_model.objects.create(plant=plant, primary_fuel=primary_fuel,
                                                num_generator=num_generator,
                                                nameplate_capacity=nameplate_capacity,
                                                num_boilers=num_boilers,year=year)
            except Exception as E:
                logger.exception('cannot creat plant',facility_code,state,latitude,longitude,E)


    def save_energy(self, generators_sheet, plant_model, plant_info_model, energy_model):
        for row in generators_sheet.iter_rows(min_row=3):            
            facility_code = row[3].value
            generator_anual_net_ = row[11].value
            if not(generator_anual_net_):
                generator_anual_net_ = decimal.Decimal(0)
            year = self.year
            try:
                plant = plant_model.objects.get(facility_code=facility_code)
                plant_information = plant_info_model.objects.get(plant=plant)
                energy = energy_model.objects.get_or_create(plant_information=plant_information,
                                                            year=year) 
                energy[0].generator_anual_net += decimal.Decimal(generator_anual_net_)
                energy[0].save()
                
            except Exception as E:
                logger.exception(E)


class Pandas_import(Import_data):
    """import data from excel file using pandas"""

    def __init__(self, filename, year):
        self.file = filename
        self.year = year
        self.year_pre = str(year)[2:]

    def process_import(self, model):
        try:
            plants_sheet = pd.read_excel(self.file, index_col=0, header=1,
                                         sheet_name="PLNT"+self.year_pre)
            plant_df = pd.DataFrame(plants_sheet,columns=['PNAME','ORISPL','PSTATABB',
                                                          'LAT','LON','NUMBLR',
                                                          'NUMGEN','PLPRMFL','NAMEPCAP'])
            plant_df = plant_df.fillna(np.nan).replace([np.nan], [None])
            generators_sheet = pd.read_excel(self.file, index_col=3, header=1, 
                                             sheet_name="GEN"+self.year_pre)
            generators_df = pd.DataFrame(plants_sheet,columns=['ORISPL','GENNTAN'])
            generators_df = generators_df.fillna(np.nan).replace([np.nan], [None])
            plant_model = model.Plant
            plant_info_model = model.Plant_information
            energy_model = model.Energy
            self.save_plants(plant_df, plant_model, plant_info_model)
            self.save_energy(generators_df, plant_model, plant_info_model, energy_model)
            logger.warning('import data done')
            return True
        except KeyError as E:
            logger.exception(E)
            return False      

    def save_plants(self, plant_df, plant_model, plant_info_model):
        for row in plant_df.itertuples():
            name = row.PNAME
            facility_code = row.ORISPL
            state = row.PSTATABB
            latitude = row.LAT
            longitude = row.LON
            primary_fuel = row.PLPRMFL
            num_generator = row.NUMGEN
            nameplate_capacity = row.NAMEPCAP
            num_boilers = row.NUMBLR
            year = self.year
            try:
                plant = plant_model.objects.get_or_create(facility_code=facility_code,
                                                          defaults={'name':name, 'state':state,
                                                                    'latitude':latitude,
                                                                    'longitude':longitude})[0]
                plant_info_model.objects.create(plant=plant, year=year,primary_fuel=primary_fuel,
                                                num_generator=num_generator,
                                                nameplate_capacity=nameplate_capacity,
                                                num_boilers=num_boilers)
            except Exception as E:
                logger.exception('cannot creat plant',plant,name,facility_code,
                                 state,latitude,longitude,primary_fuel,
                                 num_generator,nameplate_capacity,num_boilers,E)

    def save_energy(self, generators_df, plant_model, plant_info_model, energy_model):
        for row in generators_df.itertuples():            
            facility_code = row.ORISPL
            generator_anual_net_ = row.GENNTAN
            if not(generator_anual_net_):
                generator_anual_net_ = decimal.Decimal(0)
            year = self.year
            try:
                plant = plant_model.objects.get(facility_code=facility_code)
                plant_information = plant_info_model.objects.get(plant=plant, year=year)
                energy = energy_model.objects.get_or_create(plant_information=plant_information,
                                                            year=year) 
                energy[0].generator_anual_net += decimal.Decimal(generator_anual_net_)
                energy[0].save()
                
            except Exception as E:
                logger.exception('cannot create energy',facility_code,generator_anual_net_,E)
